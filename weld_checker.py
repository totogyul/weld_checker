"""
용접 로그 오매칭 검사기 — 파싱 & 비교 로직 모듈
"""

import re
import math
import io
import pdfplumber
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment


# ── 코드 정규화 ───────────────────────────────────────────────────────────────

def normalize_code(code: str) -> str:
    """숫자 세그먼트의 앞 0 제거.
    HM-004 → HM-4 / HM-0040 → HM-40 / W-04R → W-4R
    """
    parts = code.strip().upper().split('-')
    normalized = []
    for part in parts:
        m = re.fullmatch(r'(\d+)([A-Z]*)', part)
        if m:
            normalized.append(str(int(m.group(1))) + m.group(2))
        else:
            normalized.append(part)
    return '-'.join(normalized)


def is_welder_code(text: str) -> bool:
    """알파벳 2자리 - 숫자 형식 (예: HM-001, HG-02)"""
    return bool(re.fullmatch(r'[A-Z]{2}-\d+', text.strip().upper()))


def is_weld_number(text: str) -> bool:
    """W - 숫자 + 선택적 알파벳 형식 (예: W-01, W-04R)"""
    return bool(re.fullmatch(r'W-\d+[A-Z]?', text.strip().upper()))


# ── PDF 파싱 ──────────────────────────────────────────────────────────────────

def _center(word: dict) -> tuple:
    return ((word['x0'] + word['x1']) / 2, (word['top'] + word['bottom']) / 2)


def _dist(c1: tuple, c2: tuple) -> float:
    return math.sqrt((c1[0] - c2[0]) ** 2 + (c1[1] - c2[1]) ** 2)


def parse_table_pdf(pdf_file) -> list[dict]:
    """통합문서 PDF 표에서 (용접사, 용접번호) 쌍 추출.
    pdf_file: 파일 경로(str) 또는 파일류 객체
    """
    result = []
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            for row in table[1:]:  # 헤더 스킵
                if not row or len(row) < 4:
                    continue
                welder_raw = (row[2] or '').strip()
                weld_num_raw = (row[3] or '').strip()
                if is_welder_code(welder_raw) and is_weld_number(weld_num_raw):
                    result.append({
                        'welder': normalize_code(welder_raw),
                        'weld_num': normalize_code(weld_num_raw),
                        'raw_welder': welder_raw,
                        'raw_weld_num': weld_num_raw,
                    })
    return result


def parse_diagram_pdf(pdf_file, distance_threshold: float = 150) -> list[dict]:
    """도면 PDF 도형 안의 (용접사, 용접번호) 쌍 추출.
    순서 무관 — 패턴으로 구분. 가장 가까운 쌍을 greedy 매칭.
    """
    result = []
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            words = page.extract_words()
            welders   = [w for w in words if is_welder_code(w['text'])]
            weld_nums = [w for w in words if is_weld_number(w['text'])]

            paired_wn = set()

            for welder_word in welders:
                wc = _center(welder_word)
                best_dist, best_idx = float('inf'), None

                for i, wn_word in enumerate(weld_nums):
                    if i in paired_wn:
                        continue
                    d = _dist(wc, _center(wn_word))
                    if d < best_dist:
                        best_dist, best_idx = d, i

                if best_idx is not None and best_dist <= distance_threshold:
                    paired_wn.add(best_idx)
                    result.append({
                        'welder':     normalize_code(welder_word['text']),
                        'weld_num':   normalize_code(weld_nums[best_idx]['text']),
                        'raw_welder':   welder_word['text'],
                        'raw_weld_num': weld_nums[best_idx]['text'],
                    })
                else:
                    result.append({
                        'welder': normalize_code(welder_word['text']),
                        'weld_num': None,
                        'raw_welder': welder_word['text'],
                        'raw_weld_num': None,
                    })

            for i, wn_word in enumerate(weld_nums):
                if i not in paired_wn:
                    result.append({
                        'welder': None,
                        'weld_num': normalize_code(wn_word['text']),
                        'raw_welder': None,
                        'raw_weld_num': wn_word['text'],
                    })

    return result


# ── Excel 리포트 생성 ─────────────────────────────────────────────────────────

_RED    = PatternFill("solid", fgColor="FFCCCC")
_ORANGE = PatternFill("solid", fgColor="FFE5CC")
_GREEN  = PatternFill("solid", fgColor="CCFFDD")
_HDR_FILL = PatternFill("solid", fgColor="2E4D9B")
_HDR_FONT = Font(bold=True, color="FFFFFF")
_CENTER   = Alignment(horizontal='center', vertical='center')


def _style_header(ws):
    for cell in ws[1]:
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = _CENTER


def _auto_width(ws, min_w=18):
    for col in ws.columns:
        w = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max(w + 4, min_w)


def _write_pair_sheet(ws, title, pairs, fill):
    ws.title = title
    ws.append(["용접사", "용접번호"])
    _style_header(ws)
    for welder, weld_num in sorted(pairs):
        ws.append([welder, weld_num])
        for cell in ws[ws.max_row]:
            cell.fill = fill
            cell.alignment = _CENTER
    _auto_width(ws)


def build_report(table_pairs: list, diagram_pairs: list) -> bytes:
    """비교 결과를 Excel 파일로 만들어 bytes로 반환."""
    table_set   = {(p['welder'], p['weld_num']) for p in table_pairs}
    diagram_set = {
        (p['welder'], p['weld_num'])
        for p in diagram_pairs
        if p['welder'] and p['weld_num']
    }
    diagram_unpaired = [p for p in diagram_pairs if not p['welder'] or not p['weld_num']]

    only_in_diagram = diagram_set - table_set
    only_in_table   = table_set   - diagram_set
    matched         = table_set   & diagram_set

    wb = openpyxl.Workbook()

    ws1 = wb.active
    _write_pair_sheet(ws1, "도면에만있음_로그누락", only_in_diagram, _RED)

    ws2 = wb.create_sheet()
    _write_pair_sheet(ws2, "로그에만있음_도면누락", only_in_table, _ORANGE)

    ws3 = wb.create_sheet()
    _write_pair_sheet(ws3, "정상매칭", matched, _GREEN)

    ws4 = wb.create_sheet("도면_미매칭항목")
    ws4.append(["용접사(원본)", "용접번호(원본)", "비고"])
    _style_header(ws4)
    for p in diagram_unpaired:
        note = "용접번호 없음" if not p['weld_num'] else "용접사 없음"
        ws4.append([p['raw_welder'] or '-', p['raw_weld_num'] or '-', note])
    _auto_width(ws4)

    ws_sum = wb.create_sheet("요약")
    for row in [
        ("항목", "건수"),
        ("도면에만 있음 (로그 누락)", len(only_in_diagram)),
        ("로그에만 있음 (도면 누락)", len(only_in_table)),
        ("정상 매칭",                len(matched)),
        ("도면 내 미매칭 항목",       len(diagram_unpaired)),
        ("", ""),
        ("도면 총 쌍 수", len(diagram_set)),
        ("로그 총 쌍 수", len(table_set)),
    ]:
        ws_sum.append(list(row))
    _auto_width(ws_sum)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
